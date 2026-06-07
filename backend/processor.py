import os
import pandas as pd
from collections import defaultdict
from datetime import datetime
from sqlalchemy.orm import Session
from openpyxl import load_workbook as _load_workbook
from models import (
    KpiStats, StatusBreakdown, RegionBreakdown, DistrictBreakdown,
    AgeDistribution, Categorization, GenderStats, OkvedStats,
    NationalityStats, UploadSession, MicroAgg, OkvedAgg, NatAgg, NkzAgg, EduAgg, MigrationStats,
)


_CHUNK_SIZE = 30_000  # rows per chunk — keeps peak RAM under ~150 MB per chunk


def _iter_sheet_chunks(file_path: str):
    """Stream XLSX in read-only mode, yielding (sheet_name, df_chunk) up to _CHUNK_SIZE rows."""
    try:
        wb = _load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        return
    try:
        for ws in wb.worksheets:
            try:
                headers = None
                n_cols = 0
                rows: list = []
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c).strip() if c is not None else '' for c in row]
                        n_cols = len(headers)
                        continue
                    # Pad short rows so all rows have the same width
                    if len(row) < n_cols:
                        row = row + (None,) * (n_cols - len(row))
                    rows.append(row[:n_cols])
                    if len(rows) >= _CHUNK_SIZE:
                        yield ws.title, pd.DataFrame(rows, columns=headers)
                        rows = []
                if rows and headers is not None:
                    yield ws.title, pd.DataFrame(rows, columns=headers)
            except Exception:
                continue
    finally:
        wb.close()


def _update(db: Session, session_id: int, progress: int, current_file: str = ""):
    db.query(UploadSession).filter(UploadSession.id == session_id).update(
        {"progress": progress, "status": "processing", "current_file": current_file}
    )
    db.commit()


def process_excel_files(session_id: int, file_paths: list, db: Session):
    _update(db, session_id, 0)

    total_persons = 0
    active_contracts = 0
    salary_sum = 0.0
    salary_count = 0
    students = 0
    tipo_count = 0
    age_histogram: dict[int, int] = defaultdict(int)

    STATUS_KEYS = [
        "РАБОТАЮЩИЕ", "ДЕТИ ОТ 14 ДО 18 ЛЕТ", "НЕОХВАЧЕННЫЕ",
        "СТУДЕНТ", "ИП", "ПО УХОДУ ЗА РЕБЕНКОМ ДО 3",
        "ЛСИ", "ИНОСТРАННЫЕ ГРАЖДАНЕ",
        "БЕЗРАБОТНЫЕ", "БЕРЕМЕННЫЕ", "ПО УХОДУ ЗА ЛСИ",
        "ПОЛУЧАТЕЛИ АСП", "ПЕНСИОНЕРЫ", "КАНДАСЫ", "МНОГОДЕТНЫЕ", "ПОЛУЧАТЕЛИ ПОСОБИЙ",
    ]
    status_counts: dict[str, int] = {k: 0 for k in STATUS_KEYS}

    region_counts: dict[str, int] = defaultdict(int)
    region_names: dict[str, str] = {}
    district_counts: dict[tuple, int] = defaultdict(int)
    district_info: dict[tuple, dict] = {}

    age_group_ranges = [("14-17", 14, 17), ("18-24", 18, 24), ("25-29", 25, 29), ("30-35", 30, 35)]
    age_group_counts: dict[str, int] = {g: 0 for g, _, _ in age_group_ranges}

    cat_counts: dict[str, int] = defaultdict(int)
    gender_counts: dict[str, int] = defaultdict(int)
    okved_counts: dict[str, int] = defaultdict(int)
    nationality_counts: dict[str, int] = defaultdict(int)
    departed_counts: dict[str, int] = defaultdict(int)
    arrived_counts: dict[str, int] = defaultdict(int)

    total_files = len(file_paths)

    # MicroAgg: core dims only — okved/nationality live in separate tables
    _DIM_COLS = [
        "region_code", "region_name", "district_code", "district_name",
        "age_group", "age_val", "gender", "category", "family_type",
    ]
    _OKVED_DIM_COLS = ["region_code", "region_name", "district_code", "age_group", "age_val", "gender", "category", "okved"]
    _NAT_DIM_COLS   = ["region_code", "region_name", "district_code", "age_group", "age_val", "gender", "category", "nationality"]
    _NKZ_DIM_COLS   = ["region_code", "region_name", "district_code", "age_group", "age_val", "gender", "category", "nkz"]
    _EDU_DIM_COLS   = ["region_code", "region_name", "district_code", "age_group", "age_val", "gender", "category", "edu_type", "edu_name"]

    _NEW_FLAGS = ["is_asp", "is_pensioner", "is_kandas", "is_mnogodetnyi", "is_cbd"]
    _BASE_AGG = {
        "_cnt": "sum", "is_working": "sum", "is_student": "sum", "is_tipo": "sum",
        "has_contract": "sum", "is_ip": "sum", "is_lsi": "sum",
        "is_unemployed": "sum", "is_uncovered": "sum", "is_under18": "sum",
        "is_foreign": "sum", "is_pregnant": "sum", "is_uhod": "sum", "is_berkut": "sum",
        "is_asp": "sum", "is_pensioner": "sum", "is_kandas": "sum",
        "is_mnogodetnyi": "sum", "is_cbd": "sum",
        "salary": "sum", "_sal_pos": "sum", "_age_sum": "sum", "_age_pos": "sum",
    }
    _AGG_SPEC    = _BASE_AGG
    _ST_AGG_SPEC = _BASE_AGG
    _ST_RENAME = {
        "_cnt": "total_count",
        "is_working": "working_count", "is_student": "student_count",
        "is_tipo": "tipo_count", "has_contract": "contract_count",
        "is_ip": "ip_count", "is_lsi": "lsi_count",
        "is_unemployed": "unemployed_count", "is_uncovered": "uncovered_count",
        "is_under18": "under18_count", "is_foreign": "foreign_count",
        "is_pregnant": "pregnant_count", "is_uhod": "uhod_count",
        "is_berkut": "berkut_count",
        "is_asp": "asp_count", "is_pensioner": "pensioner_count",
        "is_kandas": "kandas_count", "is_mnogodetnyi": "mnogodetnyi_count",
        "is_cbd": "cbd_count",
        "salary": "salary_sum", "_sal_pos": "salary_count",
        "_age_sum": "age_sum", "_age_pos": "age_count",
    }

    for file_idx, file_path in enumerate(file_paths):
        # Reset per-file — keeps only one file's data in RAM at a time
        agg_chunks: list = []
        okved_chunks: list = []
        nat_chunks: list = []
        nkz_chunks: list = []
        edu_chunks: list = []
        file_name = os.path.basename(file_path)
        _update(db, session_id, int(file_idx / total_files * 85), file_name)

        for sheet_name, df in _iter_sheet_chunks(file_path):
            df.columns = [c.strip() if isinstance(c, str) else str(c) for c in df.columns]
            total_persons += len(df)

            def flag(col):
                if col in df.columns:
                    return pd.to_numeric(df[col], errors="coerce").fillna(0) == 1
                return pd.Series(False, index=df.index)

            opv_mask = flag("WORKING") | flag("OPV") | flag("HAS_OPV")
            ip_mask = flag("IP")
            lsi_mask = flag("LSI")
            berem_mask = flag("BEREM")
            uhod_mask = flag("DETID_DO_3LET") | flag("WOMAN_UHOD_DO3")
            berkut_mask = flag("UHOD_INV") | flag("IS_BERKUT")
            vuz_mask = flag("VUZ")
            school_mask = flag("SCHOOL")
            tipo_flag = flag("TIPO")
            asp_mask = flag("ASP")
            pensioner_mask = flag("PENSIONERS")
            kandas_mask = flag("KANDAS")
            mnogodetnyi_mask = flag("MNOGODETNYI")
            cbd_mask = flag("CBD")

            # Активный ТД: есть D_POSITION_CODE и нет TERMINATION_DATE
            if "D_POSITION_CODE" in df.columns:
                pos = df["D_POSITION_CODE"].astype(str).str.strip()
                pos_filled = ~pos.isin(["", "nan", "None", "NaT", "<NA>"])
            else:
                pos_filled = pd.Series(False, index=df.index)
            if "TERMINATION_DATE" in df.columns:
                td = df["TERMINATION_DATE"]
                term_empty = td.isna() | td.astype(str).str.strip().isin(["", "nan", "None", "NaT", "<NA>"])
            else:
                term_empty = pd.Series(True, index=df.index)
            estab_mask = pos_filled & term_empty

            status_counts["РАБОТАЮЩИЕ"] += int(opv_mask.sum())
            status_counts["ИП"] += int(ip_mask.sum())
            status_counts["ЛСИ"] += int(lsi_mask.sum())
            status_counts["БЕРЕМЕННЫЕ"] += int(berem_mask.sum())
            status_counts["ПО УХОДУ ЗА РЕБЕНКОМ ДО 3"] += int(uhod_mask.sum())
            status_counts["ПО УХОДУ ЗА ЛСИ"] += int(berkut_mask.sum())
            status_counts["ПОЛУЧАТЕЛИ АСП"] += int(asp_mask.sum())
            status_counts["ПЕНСИОНЕРЫ"] += int(pensioner_mask.sum())
            status_counts["КАНДАСЫ"] += int(kandas_mask.sum())
            status_counts["МНОГОДЕТНЫЕ"] += int(mnogodetnyi_mask.sum())
            status_counts["ПОЛУЧАТЕЛИ ПОСОБИЙ"] += int(cbd_mask.sum())
            active_contracts += int(estab_mask.sum())

            if "SMZ_3M" in df.columns:
                sal = pd.to_numeric(df["SMZ_3M"], errors="coerce")
                sal = sal[sal > 0].dropna()
                salary_sum += float(sal.sum())
                salary_count += len(sal)

            ages_s = pd.Series(0.0, index=df.index)
            has_voz = "VOZRAST" in df.columns
            age_under18 = pd.Series(False, index=df.index)
            unemployed = pd.Series(False, index=df.index)
            if has_voz:
                ages_s = pd.to_numeric(df["VOZRAST"], errors="coerce").fillna(0)
                for av, cnt in ages_s[ages_s > 0].astype(int).value_counts().items():
                    age_histogram[int(av)] += int(cnt)
                for grp, lo, hi in age_group_ranges:
                    age_group_counts[grp] += int(((ages_s >= lo) & (ages_s <= hi)).sum())

            if "DETI_DO18" in df.columns:
                age_under18 = flag("DETI_DO18")
            elif has_voz:
                age_under18 = (ages_s > 0) & (ages_s < 18)
            status_counts["ДЕТИ ОТ 14 ДО 18 ЛЕТ"] += int(age_under18.sum())

            # СТУДЕНТ = ВУЗ или ТИПО (школа не считается)
            student_mask = vuz_mask | tipo_flag
            students += int(vuz_mask.sum())       # KPI «ВУЗ» = только VUZ=1
            status_counts["СТУДЕНТ"] += int(student_mask.sum())
            tipo_count += int(tipo_flag.sum())

            if "CITIZENSHIP" in df.columns:
                foreign_mask = ~df["CITIZENSHIP"].isin(["КАЗАХСТАН", "Казахстан", "казахстан"])
                status_counts["ИНОСТРАННЫЕ ГРАЖДАНЕ"] += int(foreign_mask.sum())
            else:
                foreign_mask = pd.Series(False, index=df.index)

            has_status = (
                opv_mask | age_under18 | student_mask | tipo_flag |
                ip_mask | uhod_mask | lsi_mask | berem_mask | berkut_mask | foreign_mask
            )
            uncov = ~has_status
            status_counts["НЕОХВАЧЕННЫЕ"] += int(uncov.sum())

            if "RT_UNEMPLOYED" in df.columns:
                unemployed = flag("RT_UNEMPLOYED")
            elif has_voz:
                adult = ages_s >= 18
                unemployed = (
                    adult & ~opv_mask & ~student_mask & ~tipo_flag &
                    ~ip_mask & ~uhod_mask & ~lsi_mask & ~berem_mask & ~berkut_mask
                )
            status_counts["БЕЗРАБОТНЫЕ"] += int(unemployed.sum())

            _rn_col    = "KATO_REGNAME" if "KATO_REGNAME" in df.columns else ("REGNAME" if "REGNAME" in df.columns else None)
            _dn_col    = "KATO_RAINAME" if "KATO_RAINAME" in df.columns else ("RAINAME" if "RAINAME" in df.columns else None)
            _sdu_col   = "SDU_THZS" if "SDU_THZS" in df.columns else ("SDU_TZHS" if "SDU_TZHS" in df.columns else None)
            _okved_col = "NAME_OKED" if "NAME_OKED" in df.columns else None
            _nkz_col   = "NKZ_NAME_LVL_1" if "NKZ_NAME_LVL_1" in df.columns else ("LVL1_NAME_RU" if "LVL1_NAME_RU" in df.columns else None)

            if "KATO_REG" in df.columns and _rn_col:
                for (code, name), cnt in df.groupby(["KATO_REG", _rn_col]).size().items():
                    sc = str(code)
                    region_counts[sc] += int(cnt)
                    region_names[sc] = str(name)

            if _rn_col and _dn_col and all(c in df.columns for c in ["KATO_REG", "KATO_RAI"]):
                for (rc, rn, dc, dn), cnt in df.groupby(
                    ["KATO_REG", _rn_col, "KATO_RAI", _dn_col]
                ).size().items():
                    key = (str(rc), str(dc))
                    district_counts[key] += int(cnt)
                    district_info[key] = {
                        "reg_code": str(rc), "reg_name": str(rn),
                        "dist_code": str(dc), "dist_name": str(dn),
                    }

            if _sdu_col:
                for cat, cnt in df[_sdu_col].fillna("Не указано").value_counts().items():
                    cat_counts[str(cat)] += int(cnt)

            if "GENDER" in df.columns:
                for gen, cnt in df["GENDER"].fillna("Не указано").value_counts().items():
                    gender_counts[str(gen)] += int(cnt)

            if _okved_col:
                for nm, cnt in df[_okved_col].dropna().value_counts().items():
                    if str(nm) not in ("nan", "None"):
                        okved_counts[str(nm)] += int(cnt)
            if _nkz_col:
                for nm, cnt in df[_nkz_col].dropna().value_counts().items():
                    if str(nm) not in ("nan", "None"):
                        pass  # NkzAgg is queried directly, no separate NkzStats

            if "NATIONALTY" in df.columns:
                for nat, cnt in df["NATIONALTY"].dropna().value_counts().items():
                    nationality_counts[str(nat)] += int(cnt)

            if "A_KATO_REG_NAME" in df.columns:
                for nm, cnt in df["A_KATO_REG_NAME"].dropna().value_counts().items():
                    s = str(nm).strip()
                    if s and s not in ("nan", "None"):
                        departed_counts[s] += int(cnt)
            if "B_KATO_REG_NAME" in df.columns:
                for nm, cnt in df["B_KATO_REG_NAME"].dropna().value_counts().items():
                    s = str(nm).strip()
                    if s and s not in ("nan", "None"):
                        arrived_counts[s] += int(cnt)

            age_grp = pd.Series('', index=df.index)
            if has_voz:
                for grp, lo, hi in age_group_ranges:
                    age_grp[(ages_s >= lo) & (ages_s <= hi)] = grp

            reg_code  = df["KATO_REG"].astype(str).fillna('') if "KATO_REG" in df.columns else pd.Series('', index=df.index)
            reg_name  = df[_rn_col].astype(str).fillna('') if _rn_col else pd.Series('', index=df.index)
            dist_code = df["KATO_RAI"].astype(str).fillna('') if "KATO_RAI" in df.columns else pd.Series('', index=df.index)
            dist_name = df[_dn_col].astype(str).fillna('') if _dn_col else pd.Series('', index=df.index)
            gen_col   = df["GENDER"].fillna('Не указано').astype(str) if "GENDER" in df.columns else pd.Series('Не указано', index=df.index)
            cat_col   = df[_sdu_col].fillna('Не указано').astype(str) if _sdu_col else pd.Series('Не указано', index=df.index)
            okved_col = df[_okved_col].fillna('').astype(str) if _okved_col else pd.Series('', index=df.index)
            nkz_col   = df[_nkz_col].fillna('').astype(str) if _nkz_col else pd.Series('', index=df.index)
            nat_col   = df["NATIONALTY"].fillna('').astype(str) if "NATIONALTY" in df.columns else pd.Series('', index=df.index)
            fam_col   = df["FAMILY_TYPE"].fillna('').astype(str) if "FAMILY_TYPE" in df.columns else pd.Series('', index=df.index)
            sal_col = pd.to_numeric(df["SMZ_3M"], errors="coerce").fillna(0).clip(lower=0) if "SMZ_3M" in df.columns else pd.Series(0.0, index=df.index)
            age_int = ages_s.astype(int) if has_voz else pd.Series(0, index=df.index)

            rec_df = pd.DataFrame({
                "session_id": session_id,
                "region_code": reg_code,
                "region_name": reg_name,
                "district_code": dist_code,
                "district_name": dist_name,
                "age_group": age_grp,
                "gender": gen_col,
                "category": cat_col,
                "family_type": fam_col,
                "okved": okved_col,
                "nkz": nkz_col,
                "nationality": nat_col,
                "is_working": opv_mask.astype(int),
                "is_student": student_mask.astype(int),
                "is_tipo": tipo_flag.astype(int),
                "has_contract": estab_mask.astype(int),
                "is_ip": ip_mask.astype(int),
                "is_lsi": lsi_mask.astype(int),
                "is_unemployed": unemployed.astype(int),
                "is_uncovered": uncov.astype(int),
                "is_under18": age_under18.astype(int),
                "is_foreign": foreign_mask.astype(int),
                "is_pregnant": berem_mask.astype(int),
                "is_uhod": uhod_mask.astype(int),
                "is_berkut": berkut_mask.astype(int),
                "is_asp": asp_mask.astype(int),
                "is_pensioner": pensioner_mask.astype(int),
                "is_kandas": kandas_mask.astype(int),
                "is_mnogodetnyi": mnogodetnyi_mask.astype(int),
                "is_cbd": cbd_mask.astype(int),
                "salary": sal_col,
                "age_val": age_int,
            })

            # Clean up sentinel strings from pandas
            for col in ("region_code", "region_name", "district_code", "district_name", "gender", "category", "family_type", "okved", "nkz", "nationality", "age_group"):
                rec_df[col] = rec_df[col].replace({"nan": "", "None": ""})

            rec_df["_cnt"] = 1
            rec_df["_sal_pos"] = (rec_df["salary"] > 0).astype(int)
            rec_df["_age_pos"] = (rec_df["age_val"] > 0).astype(int)
            rec_df["_age_sum"] = rec_df["age_val"] * rec_df["_age_pos"]

            agg_chunks.append(rec_df.groupby(_DIM_COLS, as_index=False, sort=False).agg(_AGG_SPEC))

            _okved_df = rec_df[rec_df["okved"] != ""]
            if len(_okved_df):
                okved_chunks.append(_okved_df.groupby(_OKVED_DIM_COLS, as_index=False, sort=False).agg(_ST_AGG_SPEC))

            _nat_df = rec_df[rec_df["nationality"] != ""]
            if len(_nat_df):
                nat_chunks.append(_nat_df.groupby(_NAT_DIM_COLS, as_index=False, sort=False).agg(_ST_AGG_SPEC))

            _nkz_df = rec_df[rec_df["nkz"] != ""]
            if len(_nkz_df):
                nkz_chunks.append(_nkz_df.groupby(_NKZ_DIM_COLS, as_index=False, sort=False).agg(_ST_AGG_SPEC))

            vuz_name_col = df["VUZ_NAME"].fillna('').astype(str).replace({"nan": "", "None": ""}) if "VUZ_NAME" in df.columns else pd.Series('', index=df.index)
            tipo_name_col = df["TIPO_NAME"].fillna('').astype(str).replace({"nan": "", "None": ""}) if "TIPO_NAME" in df.columns else pd.Series('', index=df.index)
            school_name_col = df["SCHOOL_NAME"].fillna('').astype(str).replace({"nan": "", "None": ""}) if "SCHOOL_NAME" in df.columns else pd.Series('', index=df.index)
            edu_parts = []
            for _emask, _ecol, _etype in [
                (vuz_mask,    vuz_name_col,    "vuz"),
                (tipo_flag,   tipo_name_col,   "tipo"),
                (school_mask, school_name_col, "school"),
            ]:
                _nm = _emask & (_ecol != '') & (_ecol != 'nan') & (_ecol != 'None')
                if _nm.any():
                    _ep = rec_df[_nm].copy()
                    _ep["edu_type"] = _etype
                    _ep["edu_name"] = _ecol[_nm].values
                    edu_parts.append(_ep)
            if edu_parts:
                _edu_df = pd.concat(edu_parts, ignore_index=True)
                edu_chunks.append(_edu_df.groupby(_EDU_DIM_COLS, as_index=False, sort=False).agg(_ST_AGG_SPEC))

            del df, rec_df  # free each chunk immediately before the next

        # ── Insert this file's agg rows immediately; free RAM before next file ─
        def _insert_agg(chunks, dim_cols, agg_spec, Model):
            if not chunks:
                return
            combined = pd.concat(chunks, ignore_index=True)
            final = combined.groupby(dim_cols, as_index=False, sort=False).agg(agg_spec)
            final = final.rename(columns=_ST_RENAME)
            final["session_id"] = session_id
            final["age_val"] = final["age_val"].astype(int)
            recs = final.to_dict("records")
            for i in range(0, len(recs), 5000):
                db.bulk_insert_mappings(Model, recs[i:i + 5000])
            db.commit()

        _insert_agg(agg_chunks,   _DIM_COLS,      _AGG_SPEC,    MicroAgg)
        _insert_agg(okved_chunks, _OKVED_DIM_COLS, _ST_AGG_SPEC, OkvedAgg)
        _insert_agg(nat_chunks,   _NAT_DIM_COLS,   _ST_AGG_SPEC, NatAgg)
        _insert_agg(nkz_chunks,   _NKZ_DIM_COLS,   _ST_AGG_SPEC, NkzAgg)
        _insert_agg(edu_chunks,   _EDU_DIM_COLS,   _ST_AGG_SPEC, EduAgg)

        # Free disk immediately after this file is fully processed
        try:
            os.unlink(file_path)
        except OSError:
            pass

    # ── Global derived stats ───────────────────────────────────────────────────
    age_total = sum(age_histogram.values())
    avg_age = sum(k * v for k, v in age_histogram.items()) / age_total if age_total > 0 else 0
    cumsum = 0
    median_age = 0
    for av in sorted(age_histogram):
        cumsum += age_histogram[av]
        if cumsum >= age_total / 2:
            median_age = av
            break
    avg_salary = salary_sum / salary_count if salary_count > 0 else 0

    db.query(UploadSession).filter(UploadSession.is_active == True).update({"is_active": False})
    db.commit()

    db.add(KpiStats(
        session_id=session_id, total_persons=total_persons, total_families=0,
        working=status_counts["РАБОТАЮЩИЕ"], active_contracts=active_contracts,
        avg_salary=round(avg_salary), students=students, tipo_count=tipo_count,
        avg_age=round(avg_age, 1), median_age=round(float(median_age), 1),
    ))
    for name, count in status_counts.items():
        db.add(StatusBreakdown(session_id=session_id, status_name=name, count=count))
    for code, count in region_counts.items():
        db.add(RegionBreakdown(session_id=session_id, region_code=code,
                               region_name=region_names.get(code, code), count=count))
    for key, count in district_counts.items():
        info = district_info.get(key, {})
        db.add(DistrictBreakdown(
            session_id=session_id,
            region_code=info.get("reg_code", ""), region_name=info.get("reg_name", ""),
            district_code=info.get("dist_code", ""), district_name=info.get("dist_name", ""),
            count=count,
        ))
    for grp, lo, hi in age_group_ranges:
        db.add(AgeDistribution(session_id=session_id, age_group=grp,
                               min_age=lo, max_age=hi, count=age_group_counts[grp]))
    for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
        db.add(Categorization(session_id=session_id, category=cat, count=count))
    for gen, count in gender_counts.items():
        db.add(GenderStats(session_id=session_id, gender=gen, count=count))
    for nm, count in sorted(okved_counts.items(), key=lambda x: -x[1])[:20]:
        db.add(OkvedStats(session_id=session_id, okved_name=nm, count=count))
    for nat, count in sorted(nationality_counts.items(), key=lambda x: -x[1])[:20]:
        db.add(NationalityStats(session_id=session_id, nationality=nat, count=count))
    all_mig_regions = sorted(set(departed_counts.keys()) | set(arrived_counts.keys()))
    for reg in all_mig_regions:
        db.add(MigrationStats(
            session_id=session_id,
            region_name=reg,
            departed=departed_counts.get(reg, 0),
            arrived=arrived_counts.get(reg, 0),
        ))
    db.commit()

    db.query(UploadSession).filter(UploadSession.id == session_id).update({
        "is_active": True, "status": "done", "progress": 100,
        "total_records": total_persons,
        "completed_at": datetime.utcnow(), "current_file": "",
    })
    db.commit()
